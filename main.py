import os
import sys
import pandas as pd
import glob
from flask import Flask, request, jsonify
from flask_cors import CORS
import logging
import unicodedata
import threading
import schedule
import time
from openpyxl import load_workbook  # Importar openpyxl para leitura bruta
import re
from datetime import datetime, timedelta
import numpy as np

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Variável global para armazenar os dados
df = None

def normalize_column(col):
    """Normaliza o nome da coluna para facilitar o mapeamento"""
    if pd.isna(col):
        return ""
    col = str(col).strip().lower()
    col = ''.join(c for c in unicodedata.normalize('NFD', col) if unicodedata.category(c) != 'Mn')
    # Mapeamento direto dos nomes reais da planilha
    if col in ['cod. item', 'cod item', 'codigo item', 'cod_item']:
        return 'numero da peca'
    if col in ['descric?o', 'descrição', 'descricao', 'descri']:
        return 'descricao'
    if col in ['estoque', 'quantidade', 'qtd']:
        return 'quantidade'
    if col in ['locacao', 'locação', 'localizacao', 'localização', 'local']:
        return 'localizacao'
    # Mapeamento flexível
    col = col.replace('numero', 'num').replace('nº', 'num').replace('peca', 'peca').replace('peça', 'peca')
    col = col.replace('descrição', 'descricao').replace('descri', 'descricao')
    col = col.replace('qtd', 'quantidade').replace('quant.', 'quantidade').replace('quant', 'quantidade')
    col = col.replace('localizacao', 'localizacao').replace('localização', 'localizacao')
    return col

def load_data_from_file(file_path, file_name):
    logger.info(f"Tentando carregar {file_path}")
    df = None
    try:
        # Primeira tentativa (como no seu código original)
        df = pd.read_excel(file_path, engine='openpyxl', header=2)
    except Exception as e:
        logger.error(f"Erro ao ler {file_path} com engine padrão: {e}")
        logger.info(f"Tentando ler {file_path} sem engine específico...")
        try:
            df = pd.read_excel(file_path, engine=None, header=2) # Pandas tentará inferir
        except Exception as e_fallback:
            logger.error(f"Erro também na segunda tentativa (engine=None): {e_fallback}")
            # NOVA TENTATIVA: Ler dados brutos com openpyxl se o erro for de estilo
            if "NamedCellStyle" in str(e_fallback) and ".name should be <class 'str'> but value is <class 'NoneType'>" in str(e_fallback):
                logger.info(f"Erro de estilo detectado. Tentando leitura 'bruta' de dados com openpyxl para {file_path}")
                try:
                    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                    sheet = workbook.active # Pega a primeira aba ativa
                    # Extrai dados linha por linha
                    data_rows = []
                    for row in sheet.iter_rows():
                        data_rows.append([cell.value for cell in row])
                    if data_rows:
                        # Cria DataFrame a partir dos dados brutos (cabeçalho na primeira linha)
                        df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                        logger.info(f"Leitura 'bruta' bem-sucedida para {file_path}. {len(df)} registros.")
                    else:
                        logger.warning(f"Planilha {file_path} parece vazia na leitura 'bruta'.")
                except Exception as e_raw:
                    logger.error(f"Erro na tentativa de leitura 'bruta' com openpyxl para {file_path}: {e_raw}")
            else:
                logger.warning(f"Pulando arquivo {file_path} devido a erro não relacionado a estilo conhecido.")
    if df is not None and not df.empty:
        return df
    return None

def load_spreadsheets():
    """Carrega todas as planilhas Excel da pasta planilhas, aceitando nomes de colunas flexíveis e header na linha 3"""
    global df
    try:
        planilhas_dir = os.path.join(os.path.dirname(__file__), 'planilhas')
        logger.info(f"Procurando planilhas em: {planilhas_dir}")
        if not os.path.exists(planilhas_dir):
            logger.error(f"Diretório de planilhas não encontrado: {planilhas_dir}")
            return False
        excel_files = glob.glob(os.path.join(planilhas_dir, '*.xlsx'))
        logger.info(f"Arquivos Excel encontrados: {excel_files}")
        if not excel_files:
            logger.error("Nenhum arquivo Excel encontrado")
            return False
        all_data = []
        for file in excel_files:
            file_name = os.path.basename(file)
            temp_df = load_data_from_file(file, file_name)
            if temp_df is not None:
                # Normalizar nomes das colunas
                col_map = {}
                for col in temp_df.columns:
                    norm = normalize_column(col)
                    logger.info(f"Coluna original: '{col}' -> Normalizada: '{norm}'")
                    if 'num' in norm and 'peca' in norm:
                        col_map[col] = 'Numero da Peca'
                        logger.info(f"Mapeando '{col}' -> 'Numero da Peca'")
                    elif 'descricao' in norm:
                        col_map[col] = 'Descricao'
                        logger.info(f"Mapeando '{col}' -> 'Descricao'")
                    elif 'quantidade' in norm:
                        col_map[col] = 'Quantidade'
                        logger.info(f"Mapeando '{col}' -> 'Quantidade'")
                    elif 'localizacao' in norm:
                        col_map[col] = 'Localizacao'
                        logger.info(f"Mapeando '{col}' -> 'Localizacao'")
                temp_df = temp_df.rename(columns=col_map)
                logger.info(f"Colunas após mapeamento: {temp_df.columns.tolist()}")
                required_columns = ['Numero da Peca', 'Descricao', 'Quantidade', 'Localizacao']
                missing_columns = [col for col in required_columns if col not in temp_df.columns]
                if missing_columns:
                    logger.error(f"Colunas faltando em {file}: {missing_columns}")
                    logger.error(f"Colunas disponíveis: {temp_df.columns.tolist()}")
                    continue
                temp_df['Fonte'] = os.path.basename(file)
                temp_df['FonteData'] = extrair_data_planilha(os.path.basename(file))
                # NOVO: criar coluna normalizada já no carregamento
                temp_df['Numero da Peca Normalizado'] = temp_df['Numero da Peca'].astype(str).str.lower().str.replace(r'\s+', '', regex=True)
                for col in temp_df.columns:
                    if temp_df[col].dtype == 'object':
                        temp_df[col] = temp_df[col].astype(str).str.strip()
                all_data.append(temp_df)
                logger.info(f"Carregado {file}: {len(temp_df)} registros")
        if not all_data:
            logger.error("Nenhum dado foi carregado das planilhas")
            return False
        df = pd.concat(all_data, ignore_index=True)
        logger.info(f"Total de registros carregados: {len(df)}")
        logger.info(f"Colunas disponíveis: {df.columns.tolist()}")
        logger.info("\nExemplos de dados carregados:")
        for _, row in df.head().iterrows():
            logger.info(f"Peça: {row['Numero da Peca']}, Descrição: {row['Descricao']}")
        return True
    except Exception as e:
        logger.error(f"Erro ao carregar planilhas: {str(e)}")
        return False

def normalize_peca_num(num):
    """Remove espaços e deixa maiúsculo para normalizar número de peça"""
    if pd.isna(num):
        return ''
    return str(num).replace(' ', '').upper()

def search_stock(query):
    """Busca no estoque pelo número da peça, descrição ou localização (flexível e normalizada)"""
    if df is None or df.empty:
        logger.error("DataFrame vazio ou não inicializado")
        return []
    try:
        query_term = str(query).strip().lower().replace(' ', '')
        if not query_term:
            return []
        results_df = df.copy()
        # Busca por 'Numero da Peca Normalizado'
        if 'Numero da Peca Normalizado' in results_df.columns:
            condition_numero_peca = results_df['Numero da Peca Normalizado'].str.contains(query_term, na=False)
        else:
            condition_numero_peca = pd.Series([False] * len(results_df))
        # Busca por descrição (mantendo espaços internos)
        if 'Descricao' in results_df.columns:
            results_df['search_descricao'] = results_df['Descricao'].astype(str).str.lower()
            condition_descricao = results_df['search_descricao'].str.contains(query_term, na=False)
        else:
            condition_descricao = pd.Series([False] * len(results_df))
        # Busca por localização (sem espaços)
        if 'Localizacao' in results_df.columns:
            results_df['search_localizacao'] = results_df['Localizacao'].astype(str).str.lower().str.replace(r'\s+', '', regex=True)
            condition_localizacao = results_df['search_localizacao'].str.contains(query_term, na=False)
        else:
            condition_localizacao = pd.Series([False] * len(results_df))
        # Combina as condições (OU lógico)
        combined_condition = condition_numero_peca | condition_descricao | condition_localizacao
        results_df = results_df[combined_condition]
        # Remover duplicatas mantendo o mais recente (maior data)
        seen = {}
        for _, result in results_df.iterrows():
            key = (result['Numero da Peca'], result['Descricao'], result['Localizacao'])
            fonte_data = result.get('FonteData', '')
            try:
                data_tuple = tuple(map(int, fonte_data.split('/')[::-1]))  # (yyyy, mm, dd)
            except:
                data_tuple = (0, 0, 0)
            if key not in seen or data_tuple > seen[key][0]:
                seen[key] = (data_tuple, result)
        unique_results = [v[1].to_dict() for v in seen.values()]
        # Substituir campo Fonte pelo FonteData para exibir só a data
        for r in unique_results:
            r['Fonte'] = r.get('FonteData', r.get('Fonte', ''))
        logger.info(f"Total de resultados únicos (mais recentes): {len(unique_results)}")
        return unique_results[:20]  # Limitar a 20 resultados
    except Exception as e:
        logger.error(f"Erro durante a busca: {str(e)}")
        return []

@app.route('/stats', methods=['GET'])
def get_stats():
    """Endpoint para obter estatísticas do estoque"""
    try:
        if df is None or df.empty:
            return jsonify({
                'total_items': 0,
                'total_quantity': 0,
                'unique_locations': 0,
                'message': 'Nenhum dado carregado'
            })
        
        # Calcular estatísticas
        total_items = len(df)
        
        # Converter quantidade para numérico, tratando valores não numéricos
        try:
            df['Quantidade_Numeric'] = pd.to_numeric(df['Quantidade'], errors='coerce')
            total_quantity = int(df['Quantidade_Numeric'].sum())
        except:
            total_quantity = 0
        
        unique_locations = len(df['Localizacao'].unique())
        
        stats = {
            'total_items': total_items,
            'total_quantity': total_quantity,
            'unique_locations': unique_locations,
            'message': 'Estatísticas carregadas com sucesso'
        }
        
        logger.info(f"Estatísticas calculadas: {stats}")
        return jsonify(stats)
        
    except Exception as e:
        logger.error(f"Erro ao calcular estatísticas: {str(e)}")
        return jsonify({
            'error': str(e),
            'total_items': 0,
            'total_quantity': 0,
            'unique_locations': 0,
            'message': 'Erro ao calcular estatísticas'
        }), 500

@app.route('/search', methods=['POST'])
def search():
    """Endpoint de busca"""
    try:
        data = request.get_json()
        query = data.get('query', '').strip()
        
        logger.info(f"Recebida busca por: {query}")
        
        if not query:
            return jsonify({
                'query': '',
                'results': [],
                'total': 0,
                'message': 'Por favor, digite um termo para buscar'
            })
        
        results = search_stock(query)
        
        response = {
            'query': query,
            'results': results,
            'total': len(results),
            'message': 'Não foi encontrado essa peça nas planilhas.' if len(results) == 0 else f'Encontrados {len(results)} resultados'
        }
        
        logger.info(f"Enviando resposta: {response}")
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Erro na busca: {str(e)}")
        return jsonify({
            'error': str(e),
            'query': query if 'query' in locals() else '',
            'results': [],
            'total': 0,
            'message': 'Ocorreu um erro ao realizar a busca'
        }), 500

@app.route('/health', methods=['GET'])
def health():
    """Endpoint de verificação de saúde do sistema"""
    try:
        status = {
            'status': 'ok',
            'data_loaded': df is not None and not df.empty,
            'total_files': len(glob.glob(os.path.join(os.path.dirname(__file__), 'planilhas', '*.xlsx'))),
            'total_items': len(df) if df is not None and not df.empty else 0,
            'columns': df.columns.tolist() if df is not None and not df.empty else []
        }
        return jsonify(status)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500

@app.route('/', methods=['GET'])
def index():
    """Endpoint raiz para verificar se o servidor está funcionando"""
    return jsonify({
        'message': 'Sistema de Consulta de Estoque - API funcionando!',
        'endpoints': {
            'health': '/health',
            'stats': '/stats',
            'search': '/search (POST)'
        }
    })

def schedule_reload():
    def job():
        global df
        print('Recarregando planilhas automaticamente às 19:00...')
        load_spreadsheets()
    schedule.every().day.at('19:00').do(job)
    while True:
        schedule.run_pending()
        time.sleep(30)

def extrair_data_planilha(nome_arquivo):
    # Procura por 6 dígitos seguidos no nome do arquivo
    match = re.search(r'(\d{6})', nome_arquivo)
    if match:
        data_str = match.group(1)
        dia = data_str[:2]
        mes = data_str[2:4]
        ano = '20' + data_str[4:]  # Assume sempre 20xx
        return f"{dia}/{mes}/{ano}"
    return nome_arquivo  # fallback

def analyze_stock_changes():
    """Analisa mudanças no estoque entre diferentes datas"""
    if df is None or df.empty:
        return {"error": "Nenhum dado carregado"}
    
    try:
        # Agrupar por data e número da peça
        analysis_data = []
        
        # Obter datas únicas ordenadas
        dates = sorted(df['FonteData'].unique())
        
        for i, current_date in enumerate(dates):
            if i == 0:  # Primeira data
                continue
                
            previous_date = dates[i-1]
            
            # Dados da data atual
            current_data = df[df['FonteData'] == current_date]
            # Dados da data anterior
            previous_data = df[df['FonteData'] == previous_date]
            
            # Encontrar peças novas (não existiam na data anterior)
            current_pecas = set(current_data['Numero da Peca'].astype(str))
            previous_pecas = set(previous_data['Numero da Peca'].astype(str))
            
            new_pecas = current_pecas - previous_pecas
            removed_pecas = previous_pecas - current_pecas
            
            # Analisar mudanças de quantidade
            quantity_changes = []
            for _, current_row in current_data.iterrows():
                peca = str(current_row['Numero da Peca'])
                if peca in previous_pecas:
                    previous_row = previous_data[previous_data['Numero da Peca'].astype(str) == peca]
                    if not previous_row.empty:
                        current_qty = pd.to_numeric(current_row['Quantidade'], errors='coerce')
                        previous_qty = pd.to_numeric(previous_row.iloc[0]['Quantidade'], errors='coerce')
                        
                        if not pd.isna(current_qty) and not pd.isna(previous_qty):
                            change = current_qty - previous_qty
                            if change != 0:
                                quantity_changes.append({
                                    'peca': peca,
                                    'descricao': current_row['Descricao'],
                                    'change': change,
                                    'previous_qty': previous_qty,
                                    'current_qty': current_qty
                                })
            
            analysis_data.append({
                'date': current_date,
                'previous_date': previous_date,
                'new_items': len(new_pecas),
                'removed_items': len(removed_pecas),
                'quantity_changes': quantity_changes,
                'new_pecas_list': list(new_pecas)[:10],  # Limitar a 10 para não sobrecarregar
                'removed_pecas_list': list(removed_pecas)[:10]
            })
        
        return {
            'success': True,
            'analysis': analysis_data,
            'total_dates': len(dates),
            'date_range': f"{dates[0]} a {dates[-1]}"
        }
        
    except Exception as e:
        logger.error(f"Erro na análise de mudanças: {str(e)}")
        return {"error": str(e)}

def analyze_peca_history(peca_number):
    """Analisa o histórico de uma peça específica ao longo do tempo"""
    if df is None or df.empty:
        return {"error": "Nenhum dado carregado"}
    
    try:
        # Normalizar número da peça
        peca_normalized = str(peca_number).strip().upper()
        
        # Buscar todas as ocorrências da peça
        peca_data = df[df['Numero da Peca'].astype(str).str.contains(peca_normalized, na=False)]
        
        if peca_data.empty:
            return {"error": f"Peça {peca_number} não encontrada"}
        
        # Ordenar por data
        peca_data = peca_data.sort_values('FonteData')
        
        history = []
        for _, row in peca_data.iterrows():
            history.append({
                'date': row['FonteData'],
                'quantity': row['Quantidade'],
                'location': row['Localizacao'],
                'description': row['Descricao']
            })
        
        # Calcular estatísticas
        quantities = pd.to_numeric(peca_data['Quantidade'], errors='coerce')
        quantities = quantities.dropna()
        
        stats = {
            'total_records': len(history),
            'date_range': f"{history[0]['date']} a {history[-1]['date']}",
            'min_quantity': int(quantities.min()) if not quantities.empty else 0,
            'max_quantity': int(quantities.max()) if not quantities.empty else 0,
            'avg_quantity': float(quantities.mean()) if not quantities.empty else 0,
            'unique_locations': len(peca_data['Localizacao'].unique())
        }
        
        return {
            'success': True,
            'peca_number': peca_number,
            'history': history,
            'stats': stats
        }
        
    except Exception as e:
        logger.error(f"Erro na análise do histórico: {str(e)}")
        return {"error": str(e)}

def analyze_location_changes():
    """Analisa mudanças de localização das peças"""
    if df is None or df.empty:
        return {"error": "Nenhum dado carregado"}
    
    try:
        # Agrupar por número da peça e analisar mudanças de localização
        location_changes = []
        
        # Obter todas as peças únicas
        unique_pecas = df['Numero da Peca'].unique()
        
        for peca in unique_pecas[:100]:  # Limitar a 100 peças para performance
            peca_data = df[df['Numero da Peca'] == peca].sort_values('FonteData')
            
            if len(peca_data) > 1:
                locations = peca_data['Localizacao'].unique()
                if len(locations) > 1:
                    location_changes.append({
                        'peca': peca,
                        'description': peca_data.iloc[0]['Descricao'],
                        'locations': list(locations),
                        'first_location': peca_data.iloc[0]['Localizacao'],
                        'last_location': peca_data.iloc[-1]['Localizacao'],
                        'total_moves': len(locations) - 1
                    })
        
        return {
            'success': True,
            'location_changes': location_changes,
            'total_pecas_with_location_changes': len(location_changes)
        }
        
    except Exception as e:
        logger.error(f"Erro na análise de localizações: {str(e)}")
        return {"error": str(e)}

def get_stock_insights():
    """Gera insights gerais sobre o estoque"""
    if df is None or df.empty:
        return {"error": "Nenhum dado carregado"}
    
    try:
        insights = {
            'total_items': len(df),
            'unique_pecas': len(df['Numero da Peca'].unique()),
            'unique_locations': len(df['Localizacao'].unique()),
            'date_range': f"{df['FonteData'].min()} a {df['FonteData'].max()}",
            'total_files': len(df['Fonte'].unique())
        }
        
        # Análise de quantidade
        quantities = pd.to_numeric(df['Quantidade'], errors='coerce')
        quantities = quantities.dropna()
        
        if not quantities.empty:
            insights.update({
                'total_quantity': int(quantities.sum()),
                'avg_quantity_per_item': float(quantities.mean()),
                'items_with_zero_stock': int((quantities == 0).sum()),
                'items_with_low_stock': int((quantities <= 5).sum())
            })
        
        # Top localizações
        top_locations = df['Localizacao'].value_counts().head(5).to_dict()
        insights['top_locations'] = top_locations
        
        # Peças mais frequentes
        top_pecas = df['Numero da Peca'].value_counts().head(5).to_dict()
        insights['top_pecas'] = top_pecas
        
        return {
            'success': True,
            'insights': insights
        }
        
    except Exception as e:
        logger.error(f"Erro na geração de insights: {str(e)}")
        return {"error": str(e)}

@app.route('/analyze/changes', methods=['GET'])
def analyze_changes():
    """Endpoint para análise de mudanças no estoque"""
    try:
        result = analyze_stock_changes()
        return jsonify(result)
    except Exception as e:
        logger.error(f"Erro no endpoint de análise: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/analyze/peca/<peca_number>', methods=['GET'])
def analyze_peca(peca_number):
    """Endpoint para análise de uma peça específica"""
    try:
        result = analyze_peca_history(peca_number)
        return jsonify(result)
    except Exception as e:
        logger.error(f"Erro no endpoint de análise de peça: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/analyze/locations', methods=['GET'])
def analyze_locations():
    """Endpoint para análise de mudanças de localização"""
    try:
        result = analyze_location_changes()
        return jsonify(result)
    except Exception as e:
        logger.error(f"Erro no endpoint de análise de localizações: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/analyze/insights', methods=['GET'])
def get_insights():
    """Endpoint para insights gerais do estoque"""
    try:
        result = get_stock_insights()
        return jsonify(result)
    except Exception as e:
        logger.error(f"Erro no endpoint de insights: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/chat/analyze', methods=['POST'])
def chat_analyze():
    """Endpoint para análise inteligente via chat"""
    try:
        data = request.get_json()
        query = data.get('query', '').lower().strip()
        
        if not query:
            return jsonify({'error': 'Query não fornecida'})
        
        # Análise baseada no tipo de pergunta
        if 'mudança' in query or 'alteração' in query or 'diferença' in query:
            result = analyze_stock_changes()
            if result.get('success'):
                analysis = result['analysis']
                if analysis:
                    latest = analysis[-1]
                    response = f"📊 **Análise de Mudanças - {latest['date']}**\n\n"
                    response += f"🆕 **Novos itens:** {latest['new_items']}\n"
                    response += f"❌ **Itens removidos:** {latest['removed_items']}\n"
                    response += f"📈 **Mudanças de quantidade:** {len(latest['quantity_changes'])}\n\n"
                    
                    if latest['quantity_changes']:
                        response += "**Principais mudanças de quantidade:**\n"
                        for change in latest['quantity_changes'][:5]:
                            response += f"• {change['descricao']}: {change['previous_qty']} → {change['current_qty']} ({change['change']:+d})\n"
                    
                    return jsonify({'response': response, 'type': 'changes'})
                else:
                    return jsonify({'response': 'Não há dados suficientes para análise de mudanças.'})
            else:
                return jsonify({'response': f'Erro na análise: {result.get("error")}'})
        
        elif 'histórico' in query or 'histórico' in query:
            # Extrair número da peça da query
            import re
            peca_match = re.search(r'[A-Z]\s*\d+', query, re.IGNORECASE)
            if peca_match:
                peca_number = peca_match.group()
                result = analyze_peca_history(peca_number)
                if result.get('success'):
                    history = result['history']
                    stats = result['stats']
                    
                    response = f"📋 **Histórico da Peça {peca_number}**\n\n"
                    response += f"📅 **Período:** {stats['date_range']}\n"
                    response += f"📊 **Registros:** {stats['total_records']}\n"
                    response += f"📍 **Localizações únicas:** {stats['unique_locations']}\n\n"
                    response += f"📈 **Quantidade:** Mín: {stats['min_quantity']}, Máx: {stats['max_quantity']}, Média: {stats['avg_quantity']:.1f}\n\n"
                    
                    response += "**Histórico recente:**\n"
                    for record in history[-5:]:
                        response += f"• {record['date']}: {record['quantity']} unidades em {record['location']}\n"
                    
                    return jsonify({'response': response, 'type': 'history'})
                else:
                    return jsonify({'response': f'Erro na análise: {result.get("error")}'})
            else:
                return jsonify({'response': 'Por favor, especifique o número da peça para análise do histórico.'})
        
        elif 'localização' in query or 'localizacao' in query:
            result = analyze_location_changes()
            if result.get('success'):
                changes = result['location_changes']
                response = f"📍 **Análise de Mudanças de Localização**\n\n"
                response += f"📊 **Peças com mudanças:** {result['total_pecas_with_location_changes']}\n\n"
                
                if changes:
                    response += "**Principais mudanças:**\n"
                    for change in changes[:5]:
                        response += f"• {change['peca']} ({change['description']}): {change['first_location']} → {change['last_location']}\n"
                
                return jsonify({'response': response, 'type': 'locations'})
            else:
                return jsonify({'response': f'Erro na análise: {result.get("error")}'})
        
        elif 'insights' in query or 'resumo' in query or 'estatísticas' in query:
            result = get_stock_insights()
            if result.get('success'):
                insights = result['insights']
                
                response = f"📊 **Insights do Estoque**\n\n"
                response += f"📦 **Total de itens:** {insights['total_items']:,}\n"
                response += f"🔢 **Peças únicas:** {insights['unique_pecas']:,}\n"
                response += f"📍 **Localizações únicas:** {insights['unique_locations']:,}\n"
                response += f"📅 **Período:** {insights['date_range']}\n"
                response += f"📁 **Arquivos:** {insights['total_files']}\n\n"
                
                if 'total_quantity' in insights:
                    response += f"📈 **Quantidade total:** {insights['total_quantity']:,}\n"
                    response += f"📊 **Média por item:** {insights['avg_quantity_per_item']:.1f}\n"
                    response += f"⚠️ **Itens sem estoque:** {insights['items_with_zero_stock']:,}\n"
                    response += f"🔴 **Estoque baixo (≤5):** {insights['items_with_low_stock']:,}\n\n"
                
                response += "**Top 5 Localizações:**\n"
                for loc, count in insights['top_locations'].items():
                    response += f"• {loc}: {count} itens\n"
                
                return jsonify({'response': response, 'type': 'insights'})
            else:
                return jsonify({'response': f'Erro na análise: {result.get("error")}'})
        
        else:
            return jsonify({'response': 'Posso ajudar com:\n• Análise de mudanças no estoque\n• Histórico de peças específicas\n• Mudanças de localização\n• Insights gerais\n\nTente perguntar sobre "mudanças", "histórico da peça X", "localizações" ou "insights".'})
        
    except Exception as e:
        logger.error(f"Erro no chat analyze: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Carregar dados ao iniciar2
    if load_spreadsheets():
        logger.info("Sistema iniciado com sucesso!")
        logger.info("Servidor rodando em: http://localhost:5000")
        logger.info("Endpoints disponíveis:")
        logger.info("  - GET  / (informações)")
        logger.info("  - GET  /health (status do sistema)")
        logger.info("  - GET  /stats (estatísticas)")
        logger.info("  - POST /search (busca no estoque)")
    else:
        logger.error("Erro ao carregar dados. Verifique os logs acima.")
    # Iniciar agendamento em thread separada2
    t = threading.Thread(target=schedule_reload, daemon=True)
    t.start()
    app.run(host='0.0.0.0', port=5000, debug=True)